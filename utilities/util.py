"""
@package utilities

Util class implementation
All most commonly used utilities should be implemented in this class

Example:
    name = self.util.getUniqueName()
"""
import time
import traceback
import random, string
import openpyxl
import utilities.custom_logger as cl
import logging


class Util(object):

    log = cl.customLogger(logging.INFO)

    def sleep(self, sec, info=""):
        """
        Put the program to wait for the specified amount of time
        """
        if info is not None:
            self.log.info("Wait :: '" + str(sec) + "' seconds for " + info)
        try:
            time.sleep(sec)
        except InterruptedError:
            traceback.print_stack()

    def get_alpha_numeric(self, length, type='letters'):
        """
        Get random string of characters

        Parameters:
            length: Length of string, number of characters string should have
            type: Type of characters string should have. Default is letters
            Provide lower/upper/digits for different types
        """
        alpha_num = ''
        if type == 'lower':
            case = string.ascii_lowercase
        elif type == 'upper':
            case = string.ascii_uppercase
        elif type == 'digits':
            case = string.digits
        elif type == 'mix':
            case = string.ascii_letters + string.digits
        else:
            case = string.ascii_letters
        return alpha_num.join(random.choice(case) for i in range(length))

    def get_unique_name(self, char_count=10):
        """
        Get a unique name
        """
        return self.get_alpha_numeric(char_count, 'lower')

    def get_unique_name_list(self, list_size=5, item_length=None):
        """
        Get a list of valid email ids

        Parameters:
            list_size: Number of names. Default is 5 names in a list
            item_length: It should be a list containing number of items equal to the listSize
                        This determines the length of the each item in the list -> [1, 2, 3, 4, 5]
        """
        name_list = []
        for i in range(0, list_size):
            name_list.append(self.get_unique_name(item_length[i]))
        return name_list

    def verify_list_match(self, expected_list, actual_list):
        """
        Verify two list matches

        Parameters:
            expected_list: Expected List
            actual_list: Actual List
        """
        return set(expected_list) == set(actual_list)

    def verify_list_contains(self, expected_list, actual_list):
        """
        Verify actual list contains elements of expected list

        Parameters:
            expected_list: Expected List
            actual_list: Actual List
        """
        length = len(expected_list)
        for i in range(0, length):
            if expected_list[i] not in actual_list:
                return False
        else:
            return True

    def read_data(self, search_name="", row="", col=""):
        path = "C:\\Users\\Sagar\\Desktop\\Book1.xlsx"

        # workbook object is created
        wb_obj = openpyxl.load_workbook(path)

        sheet_obj = wb_obj.active

        # print total number of column
        print(sheet_obj.max_column)
        print(sheet_obj.max_row)

        m_row = sheet_obj.max_row
        max_col = sheet_obj.max_column
        name = search_name

        print("******************ROW PRINT********************************")

        for i in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=row, column=i)
            print(cell_obj.value)

        print("**********************COLUMN PRINT****************************")

        for j in range(2, m_row + 1):
            cell_obj = sheet_obj.cell(row=j, column=col)
            print(cell_obj.value)

        print("**************************************************")

        for i in range(1, m_row + 1):
            for j in range(1, max_col + 1):

                cell_obj = sheet_obj.cell(row=i, column=j)
                if cell_obj.value == name:
                    print(cell_obj.value)
                    break
