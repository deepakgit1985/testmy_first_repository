import openpyxl


def get_row_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet_obj = workbook.active
    return sheet_obj.max_row


def get_column_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet_obj = workbook.active
    return sheet_obj.max_column


def read_data(file, row_num, col_num):
    workbook = openpyxl.load_workbook(file)
    sheet_obj = workbook.active
    return sheet_obj.cell(row_num, col_num).value


def write_data(file, row_num, col_num, data):
    workbook = openpyxl.load_workbook(file)
    sheet_obj = workbook.active
    sheet_obj.cell(row_num, col_num).value = data
    workbook.save(file)


def fill_green_color(file, row_num, col_num):
    workbook = openpyxl.load_workbook(file)
    sheet_obj = workbook.active
    green_fill = PatternFill(start_color='60b212',
                             end_color='60b212',
                             fill_type="solid")
    sheet_obj.cell(row_num, col_num).fill = green_fill
    workbook.save(file)
